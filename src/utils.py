#!/usr/bin/python3

# MIT License
# 
# Copyright (c) 2017 Josh Simonot
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
# 

import calendar
import datetime
from openpyxl import load_workbook

#==========================
def add_months(sourcedate, num_months):
	month = sourcedate.month - 1 + num_months
	year = int(sourcedate.year + month / 12 )
	month = month % 12 + 1
	day = min(sourcedate.day,calendar.monthrange(year,month)[1])
	return datetime.date(year,month,day)

#==========================
def dateNotInRange( startDate, testDate, endDate ):
	if startDate.year < testDate.year < endDate.year:
		return False
	
	if startDate.year == testDate.year and startDate.month <= testDate.month:
		return False
		
	if endDate.year == testDate.year and testDate.month <= endDate.month:
		return False
		
	return True

#==========================
def dateInRange( startDate, testDate, endDate ):
	return not dateNotInRange( startDate, testDate, endDate )
	
#==========================
def loadDataFromExcel(filepath):	
	wb = load_workbook(filename=filepath, read_only=True, data_only=True)
	ws_info = wb["Info"]
	
	data = {}
	
	#get date range
	dateRange = []
	data["dateRange"] = dateRange
	
	startDate = ws_info["E3"].value
	endDate = ws_info["E4"].value
	
	dateRange.append(startDate)
	tmpdate = add_months(startDate, 1)
	
	while dateInRange( startDate, tmpdate, endDate ):
		dateRange.append( tmpdate )
		tmpdate = add_months(tmpdate, 1)
	
	#get account names (and init data structures)
	accounts = {}
	data["accounts"] = accounts
	
	column = "B{}"
	row = 3
	while True:
		cell = column.format(row)
		accountName = ws_info[cell].value
		if accountName == None: 
			break #stop loop
		else:
			accounts[accountName] = {} #create data structure
			accounts[accountName]["row"] = 2*row-3
			accounts[accountName]["value"] = []
			accounts[accountName]["cw"] = []
			
		row = row + 1
	
	#for each year (worksheets)
	for year in range( startDate.year, endDate.year+1 ):
		ws_year = wb[ str(year) ] 
		
		#for each valid month
		for column in range(3, 3+12):
			date = ws_year.cell(row=2, column=column).value
			if dateNotInRange( startDate, date, endDate ):
				continue #next loop iteration (ie: skip this date)

			#collect data (account value and contributions/withdrawals) for each account
			for accountName,accountData in accounts.items():					
				row = accountData["row"]
				
				value = ws_year.cell( row=row, column=column ).value
				accountData["value"].append( 0 if value==None else value )
				
				cw = ws_year.cell( row=(row+1), column=column ).value
				accountData["cw"].append( 0 if cw==None else cw )
	
	return data
	
#==========================
def builddashboard(filename, plot1, plot2, plot3, plot4, plot5, bgcolor_hex):
	htmltext = """
<!DOCTYPE html>
<html>
	<head>
		<style>
		.chart-stage {{
			display: inline-block;
			margin: 0.3%;
			border: 1px solid #444444; 
		}}
		</style>
	</head>
		
	<body style="background-color:{};">
		<div class="chart-stage" style="width:49%;">
		<iframe width="100%" height="525px" frameborder="0" scrolling="no" src="{}"></iframe>
		</div>
		
		<div class="chart-stage" style="width:49%;">
		<iframe width="100%" height="525px" frameborder="0" scrolling="no" src="{}"></iframe>
		</div>
				
		<div class="chart-stage" style="width:99%;">
		<iframe width="100%" height="525" frameborder="0" scrolling="no" src="{}"></iframe>
		</div>
		<div class="chart-stage" style="width:99%;">
		<iframe width="100%" height="525" frameborder="0" scrolling="no" src="{}"></iframe>
		</div>
		<div class="chart-stage" style="width:99%;">
		<iframe width="100%" height="525" frameborder="0" scrolling="no" src="{}"></iframe>
		</div>
	</body>
</html>
"""
	htmltext = htmltext.format( bgcolor_hex, plot1, plot2, plot3, plot4, plot5 )
	with open(filename, 'w') as file:
		file.write(htmltext)
		file.close()
		
	return filename
#==========================









